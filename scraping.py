import openpyxl as xl
from flask import Flask, render_template, request

# Variable declarations
matched_ingredients_counter = 0
dishes = []
match_rate = 1
the_bar = 100

# User's selected items
selected_items = ["beef", "milk", "poppy", "lemon"]

# Load excel file
wb = xl.load_workbook("Recipes_Better_3.xlsx")
sheet = wb["Sheet1"]


# Main loop that will keep running until the lists has all (limit of 12) possible dishes
while len(dishes) <= 33 or match_rate > 0:

    # Nested loop to reach each cell and save it
    for row in range(1, sheet.max_row):
        for column in range(2, 19):
            cell = sheet.cell(row, column)

            # Cycle through every ingredient
            for item in selected_items:

                # Execute if the ingredient is in that dish
                if item in str(cell.value):

                    # Calculate the match rate between dish ingreds. and user ingreds.
                    matched_ingredients_counter += 1
                    match_rate = (matched_ingredients_counter / len(selected_items)) * 100

                    # If the match rate meets the minimum and isn't in the array, append all relevant info
                    if match_rate > the_bar and not sheet.cell(row, 1).value in dishes and not sheet.cell(row, 20).value in dishes and not sheet.cell(row, 21).value in dishes:
                        dishes.append(sheet.cell(row, 1).value)
                        dishes.append(sheet.cell(row, 20).value)
                        dishes.append(sheet.cell(row, 21).value)

                # No matches found in the dish
                else:
                    match_rate = 0
        # Lower the bar each run, to get best results
        matched_ingredients_counter = 0
    the_bar -= 20

print(dishes)

